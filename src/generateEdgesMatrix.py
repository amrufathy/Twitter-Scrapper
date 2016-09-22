import os
from openpyxl import Workbook, load_workbook

# -----------------------------------------------------------------------
# load our API credentials
# -----------------------------------------------------------------------
settings = {}
execfile("settings.py", settings)

# -----------------------------------------------------------------------
# make the nodes list
# -----------------------------------------------------------------------
nodes = {}
execfile("generateNodeMatrix.py", nodes)
nodes = nodes["matrix"]

files = []

matrix = {}

for output_file in os.listdir(settings["files_directory_name"]):
    files.append(os.path.splitext(output_file)[0])

for file in files:
    fileName = settings["files_directory_name"] + str(file) + '.xlsx'

    workbook = load_workbook(fileName)

    try:
        worksheet = workbook["followingList"]  # user follows who, user is the source
    except:
        pass

    if file not in matrix.keys():
        matrix[file] = []

    for entry in worksheet["D"][1:]:  # 'D' column has screen names
        matrix[file].append(entry.value)

    try:
        worksheet = workbook["followersList"]  # who follows user, user is the target
    except:
        pass

    for entry in worksheet["D"][1:]:  # 'D' column has screen names
        if entry.value not in matrix.keys():
            matrix[entry.value] = [file]
        else:
            matrix[entry.value].append(file)

# --------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------

# generate matrix file

workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'edges'

worksheet['A1'] = 'id'
worksheet['B1'] = 'source'
worksheet['C1'] = 'target'
worksheet['D1'] = 'source id'
worksheet['E1'] = 'target id'

j = 2

for key in matrix:
    for value in matrix[key]:
        worksheet['A' + str(j)] = j - 1
        worksheet['B' + str(j)] = str(key)
        worksheet['C' + str(j)] = str(value)
        worksheet['D' + str(j)] = (nodes.index(str(key)) + 1)
        worksheet['E' + str(j)] = (nodes.index(str(value)) + 1)
        j += 1

workbook.save(settings["matrix_directory_name"] + "edges.xlsx")
