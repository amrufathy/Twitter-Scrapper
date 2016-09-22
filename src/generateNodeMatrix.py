import os
from openpyxl import Workbook, load_workbook

# -----------------------------------------------------------------------
# load our API credentials
# -----------------------------------------------------------------------
settings = {}
execfile("settings.py", settings)

files = []

matrix = []

for output_file in os.listdir(settings["files_directory_name"]):
    files.append(os.path.splitext(output_file)[0])

for file in files:
    fileName = settings["files_directory_name"] + str(file) + '.xlsx'

    workbook = load_workbook(fileName)

    try:
        worksheet = workbook["followingList"]  # user follows who, user is the source
    except:
        pass

    if file not in matrix:
        matrix.append(file)

    for entry in worksheet["D"][1:]:  # 'D' column has screen names
        if entry.value not in matrix:
            matrix.append(entry.value)

    try:
        worksheet = workbook["followersList"]  # who follows user, user is the target
    except:
        pass

    for entry in worksheet["D"][1:]:  # 'D' column has screen names
        if entry.value not in matrix:
            matrix.append(entry.value)

# --------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------

# generate matrix file

workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'nodes'

worksheet['A1'] = 'id'
worksheet['B1'] = 'name'

j = 2

for key in matrix:
    worksheet['A' + str(j)] = j - 1
    worksheet['B' + str(j)] = str(key)
    j += 1

workbook.save(settings["matrix_directory_name"] + "nodes.xlsx")
